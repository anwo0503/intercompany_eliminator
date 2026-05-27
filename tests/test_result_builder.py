from __future__ import annotations

import os
import tempfile
from datetime import date

import pandas as pd
import pytest

from engine.matcher import MatchResult
from openpyxl import load_workbook

from engine.result_builder import (
    COLUMNS,
    _LABEL_MISMATCH,
    _LABEL_UNMATCHED,
    _SUBTOTAL_ITEM_LABEL,
    build_result_table,
    export_to_excel,
)

_MATCH_COLS = [
    "buy_item", "buy_po_code", "buy_invoice_date", "buy_quantity", "buy_total_money",
    "sell_item", "sell_so_code", "sell_invoice_date", "sell_quantity", "sell_total_money",
]
_BUY_COLS = ["item", "po_code", "invoice_date", "quantity", "total_money"]
_SELL_COLS = ["item", "so_code", "invoice_date", "quantity", "total_money"]


def _match_row(
    buy_item="ItemA",
    buy_po="PO-1",
    buy_date=date(2024, 1, 1),
    buy_qty=10,
    buy_money=1000,
    sell_item="ItemA",
    sell_so="SO-1",
    sell_date=date(2024, 1, 1),
    sell_qty=10,
    sell_money=1000,
) -> dict:
    return {
        "buy_item": buy_item, "buy_po_code": buy_po, "buy_invoice_date": buy_date,
        "buy_quantity": buy_qty, "buy_total_money": buy_money,
        "sell_item": sell_item, "sell_so_code": sell_so, "sell_invoice_date": sell_date,
        "sell_quantity": sell_qty, "sell_total_money": sell_money,
    }


def _make_result(exact=None, mismatches=None, unmatched_buy=None, unmatched_sell=None) -> MatchResult:
    return MatchResult(
        exact_matches=pd.DataFrame(exact or [], columns=_MATCH_COLS),
        amount_mismatches=pd.DataFrame(mismatches or [], columns=_MATCH_COLS),
        unmatched_buy=pd.DataFrame(unmatched_buy or [], columns=_BUY_COLS),
        unmatched_sell=pd.DataFrame(unmatched_sell or [], columns=_SELL_COLS),
    )


def _section_indices(df: pd.DataFrame) -> tuple[int, int]:
    """Return (mismatch_label_idx, unmatched_label_idx) in the result DataFrame."""
    buyer_col = df["Buyer Side"].tolist()
    return buyer_col.index(_LABEL_MISMATCH), buyer_col.index(_LABEL_UNMATCHED)


def test_all_three_sections_in_correct_order():
    """2 exact matches + 1 amount mismatch + 1 unmatched produce all three sections in order."""
    result = _make_result(
        exact=[
            _match_row(buy_item="Beta", buy_date=date(2024, 1, 2)),
            _match_row(buy_item="Alpha", buy_date=date(2024, 1, 1)),
        ],
        mismatches=[
            _match_row(buy_item="Gamma", buy_money=1000, sell_money=1005),
        ],
        unmatched_buy=[
            {"item": "Delta", "po_code": "PO-X", "invoice_date": date(2024, 1, 3), "quantity": 5, "total_money": 500},
        ],
    )
    df = build_result_table(result, "BUY_CO", "SELL_CO")

    mismatch_idx, unmatched_idx = _section_indices(df)

    # Section 1 data exists before the mismatch separator
    assert mismatch_idx >= 2  # at least 2 exact match rows before sep
    # Sections appear in the right order
    assert mismatch_idx < unmatched_idx
    # Section 1 data rows have expected Buyer Side value
    assert df.iloc[0]["Buyer Side"] == "BUY_CO"
    # Section 2 data exists between the two separators
    assert unmatched_idx > mismatch_idx + 1
    # Section 3 data exists after the unmatched separator
    assert len(df) > unmatched_idx + 1


def test_section1_sorted_by_item_then_date():
    """Rows in section 1 are sorted by Item ascending, then invoice date ascending."""
    result = _make_result(
        exact=[
            _match_row(buy_item="Beta", buy_date=date(2024, 1, 2)),
            _match_row(buy_item="Alpha", buy_date=date(2024, 1, 3)),
            _match_row(buy_item="Alpha", buy_date=date(2024, 1, 1)),
        ],
    )
    df = build_result_table(result, "B", "S")

    mismatch_idx, _ = _section_indices(df)
    # blank row at mismatch_idx-1, subtotal at mismatch_idx-2; data rows before that
    section1_data = df.iloc[: mismatch_idx - 2].reset_index(drop=True)

    assert section1_data["Item"].tolist() == ["Alpha", "Alpha", "Beta"]
    alpha = section1_data[section1_data["Item"] == "Alpha"]["Invoice date — Buyer side"].tolist()
    assert alpha == sorted(alpha)


def test_unmatched_buy_blank_columns():
    """Unmatched buy row: columns D (Seller date), H (Seller money), K (SO) are blank."""
    result = _make_result(
        unmatched_buy=[
            {"item": "ItemX", "po_code": "PO-1", "invoice_date": date(2024, 3, 1), "quantity": 2, "total_money": 200},
        ],
    )
    df = build_result_table(result, "B", "S")

    _, unmatched_idx = _section_indices(df)
    section3 = df.iloc[unmatched_idx + 1 :].reset_index(drop=True)
    row = section3.iloc[0]

    # Blank: D, H, K
    assert pd.isna(row["Invoice date — Seller side"])
    assert pd.isna(row["Total money — Seller side"])
    assert pd.isna(row["Sales order"])
    # Present: C, G, J
    assert row["Invoice date — Buyer side"] == date(2024, 3, 1)
    assert row["Total money — Buyer side"] == 200
    assert row["Purchase order"] == "PO-1"


def test_unmatched_sell_blank_columns():
    """Unmatched sell row: columns C (Buyer date), G (Buyer money), J (PO) are blank."""
    result = _make_result(
        unmatched_sell=[
            {"item": "ItemY", "so_code": "SO-2", "invoice_date": date(2024, 4, 1), "quantity": 3, "total_money": 300},
        ],
    )
    df = build_result_table(result, "B", "S")

    _, unmatched_idx = _section_indices(df)
    section3 = df.iloc[unmatched_idx + 1 :].reset_index(drop=True)
    row = section3.iloc[0]

    # Blank: C, G, J
    assert pd.isna(row["Invoice date — Buyer side"])
    assert pd.isna(row["Total money — Buyer side"])
    assert pd.isna(row["Purchase order"])
    # Present: D, H, K
    assert row["Invoice date — Seller side"] == date(2024, 4, 1)
    assert row["Total money — Seller side"] == 300
    assert row["Sales order"] == "SO-2"


def test_separator_rows_with_blank_preceding_row():
    """Section separator rows exist and are preceded by a blank row."""
    result = _make_result(
        exact=[_match_row()],
        mismatches=[_match_row(buy_money=1000, sell_money=1005)],
        unmatched_buy=[
            {"item": "X", "po_code": "PO-Z", "invoice_date": None, "quantity": 1, "total_money": 100},
        ],
    )
    df = build_result_table(result, "B", "S")

    mismatch_idx, unmatched_idx = _section_indices(df)

    # Blank row immediately before each label
    assert pd.isna(df.iloc[mismatch_idx - 1]["Buyer Side"])
    assert pd.isna(df.iloc[unmatched_idx - 1]["Buyer Side"])


def test_export_to_excel_creates_readable_file():
    """export_to_excel writes a valid Excel file with the correct number of columns."""
    result = _make_result(
        exact=[_match_row(buy_date=date(2024, 1, 1), sell_date=date(2024, 1, 1))],
    )
    df = build_result_table(result, "BUY", "SELL")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        from engine.translations import COL_KEY, set_language, t
        set_language("en")
        export_to_excel(df, path)
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        wb_df = pd.read_excel(path)
        expected_cols = [t(COL_KEY[col]) for col in COLUMNS]
        assert list(wb_df.columns) == expected_cols
    finally:
        os.unlink(path)


def test_separator_black_font_in_export():
    """Separator rows in the exported Excel file have black (FF000000) font color."""
    result = _make_result(
        exact=[_match_row()],
        mismatches=[_match_row(buy_money=1000, sell_money=1005)],
    )
    df = build_result_table(result, "B", "S")

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        export_to_excel(df, path)
        wb = load_workbook(path)
        ws = wb.active
        for row_idx in range(2, ws.max_row + 1):
            buyer_cell = ws.cell(row=row_idx, column=1)
            if buyer_cell.value in (_LABEL_MISMATCH, _LABEL_UNMATCHED):
                for cell in ws[row_idx]:
                    assert cell.font.color.rgb == "FF000000", (
                        f"Separator row {row_idx} has non-black font: {cell.font.color.rgb}"
                    )
    finally:
        os.unlink(path)


def test_unmatched_buy_seller_side_blank():
    """Unmatched buy-side rows have a blank Seller Side column."""
    result = _make_result(
        unmatched_buy=[
            {"item": "ItemX", "po_code": "PO-1", "invoice_date": date(2024, 3, 1), "quantity": 2, "total_money": 200},
        ],
    )
    df = build_result_table(result, "B", "S")

    _, unmatched_idx = _section_indices(df)
    section3 = df.iloc[unmatched_idx + 1:].reset_index(drop=True)
    buy_row = section3[section3["Buyer Side"] == "B"].iloc[0]

    assert pd.isna(buy_row["Seller Side"])


def test_unmatched_sell_buyer_side_blank():
    """Unmatched sell-side rows have a blank Buyer Side column."""
    result = _make_result(
        unmatched_sell=[
            {"item": "ItemY", "so_code": "SO-2", "invoice_date": date(2024, 4, 1), "quantity": 3, "total_money": 300},
        ],
    )
    df = build_result_table(result, "B", "S")

    _, unmatched_idx = _section_indices(df)
    section3 = df.iloc[unmatched_idx + 1:].reset_index(drop=True)
    sell_row = section3[section3["Seller Side"] == "S"].iloc[0]

    assert pd.isna(sell_row["Buyer Side"])


def test_subtotal_rows_exist_with_correct_sums():
    """Each non-empty section ends with a subtotal row containing correct column sums."""
    result = _make_result(
        exact=[
            _match_row(buy_item="A", buy_qty=5, buy_money=500, sell_money=500),
            _match_row(buy_item="B", buy_qty=3, buy_money=300, sell_money=300),
        ],
        mismatches=[
            _match_row(buy_item="C", buy_qty=2, buy_money=200, sell_money=195),
        ],
        unmatched_buy=[
            {"item": "D", "po_code": "PO-D", "invoice_date": date(2024, 5, 1), "quantity": 4, "total_money": 400},
        ],
        unmatched_sell=[
            {"item": "E", "so_code": "SO-E", "invoice_date": date(2024, 5, 2), "quantity": 6, "total_money": 600},
        ],
    )
    df = build_result_table(result, "B", "S")

    mismatch_idx, unmatched_idx = _section_indices(df)

    # Section 1 subtotal is the row just before the blank+separator
    s1_subtotal = df.iloc[mismatch_idx - 2]
    assert s1_subtotal["Item"] == _SUBTOTAL_ITEM_LABEL
    assert s1_subtotal["Quantity"] == 8   # 5 + 3
    assert s1_subtotal["Total money — Buyer side"] == 800   # 500 + 300
    assert s1_subtotal["Total money — Seller side"] == 800  # 500 + 300
    assert s1_subtotal["Difference"] == 0  # 0 + 0

    # Section 2 subtotal
    s2_subtotal = df.iloc[unmatched_idx - 2]
    assert s2_subtotal["Item"] == _SUBTOTAL_ITEM_LABEL
    assert s2_subtotal["Quantity"] == 2
    assert s2_subtotal["Total money — Buyer side"] == 200
    assert s2_subtotal["Total money — Seller side"] == 195
    assert s2_subtotal["Difference"] == 5  # 200 - 195

    # Section 3 subtotal is the last row
    s3_subtotal = df.iloc[-1]
    assert s3_subtotal["Item"] == _SUBTOTAL_ITEM_LABEL
    assert s3_subtotal["Quantity"] == 10   # 4 + 6
    # Buy-side money: only the buy unmatched row (400), sell side is None
    assert s3_subtotal["Total money — Buyer side"] == 400
    # Sell-side money: only the sell unmatched row (600)
    assert s3_subtotal["Total money — Seller side"] == 600
    # Difference: all None in unmatched section
    assert pd.isna(s3_subtotal["Difference"])

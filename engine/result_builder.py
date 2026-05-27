from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from engine.matcher import MatchResult
from engine.translations import COL_KEY, SEP_KEY, t

COLUMNS = [
    "Buyer Side",
    "Seller Side",
    "Invoice date — Buyer side",
    "Invoice date — Seller side",
    "Item",
    "Quantity",
    "Total money — Buyer side",
    "Total money — Seller side",
    "Difference",
    "Purchase order",
    "Sales order",
]

_LABEL_MISMATCH = "--- AMOUNT MISMATCHES ---"
_LABEL_UNMATCHED = "--- UNMATCHED TRANSACTIONS ---"
_SEPARATOR_LABELS = {_LABEL_MISMATCH, _LABEL_UNMATCHED}

_SUBTOTAL_ITEM_LABEL = "Subtotal"

_DATE_COLS = {"Invoice date — Buyer side", "Invoice date — Seller side"}
_MONEY_COLS = {"Total money — Buyer side", "Total money — Seller side", "Difference"}

_MONEY_FORMAT = "#,##0"
_DATE_FORMAT = "DD/MM/YYYY"


def _matched_rows(df: pd.DataFrame, buyer_short: str, seller_short: str) -> list[dict]:
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "Buyer Side": buyer_short,
            "Seller Side": seller_short,
            "Invoice date — Buyer side": r["buy_invoice_date"],
            "Invoice date — Seller side": r["sell_invoice_date"],
            "Item": r["buy_item"],
            "Quantity": int(r["buy_quantity"]),
            "Total money — Buyer side": int(r["buy_total_money"]),
            "Total money — Seller side": int(r["sell_total_money"]),
            "Difference": int(r["buy_total_money"]) - int(r["sell_total_money"]),
            "Purchase order": r["buy_po_code"],
            "Sales order": r["sell_so_code"],
        })
    return rows


def _unmatched_buy_rows(df: pd.DataFrame, buyer_short: str, seller_short: str) -> list[dict]:
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "Buyer Side": buyer_short,
            "Seller Side": None,
            "Invoice date — Buyer side": r["invoice_date"],
            "Invoice date — Seller side": None,
            "Item": r["item"],
            "Quantity": int(r["quantity"]),
            "Total money — Buyer side": int(r["total_money"]),
            "Total money — Seller side": None,
            "Difference": None,
            "Purchase order": r["po_code"],
            "Sales order": None,
        })
    return rows


def _unmatched_sell_rows(df: pd.DataFrame, buyer_short: str, seller_short: str) -> list[dict]:
    rows = []
    for _, r in df.iterrows():
        rows.append({
            "Buyer Side": None,
            "Seller Side": seller_short,
            "Invoice date — Buyer side": None,
            "Invoice date — Seller side": r["invoice_date"],
            "Item": r["item"],
            "Quantity": int(r["quantity"]),
            "Total money — Buyer side": None,
            "Total money — Seller side": int(r["total_money"]),
            "Difference": None,
            "Purchase order": None,
            "Sales order": r["so_code"],
        })
    return rows


def _blank_row() -> dict:
    return {col: None for col in COLUMNS}


def _label_row(label: str) -> dict:
    row = _blank_row()
    row["Buyer Side"] = label
    return row


def _subtotal_row(section_df: pd.DataFrame) -> dict:
    row = _blank_row()
    row["Item"] = _SUBTOTAL_ITEM_LABEL

    def _sum_col(col: str) -> int | None:
        vals = section_df[col].dropna()
        return int(vals.sum()) if len(vals) > 0 else None

    row["Quantity"] = _sum_col("Quantity")
    row["Total money — Buyer side"] = _sum_col("Total money — Buyer side")
    row["Total money — Seller side"] = _sum_col("Total money — Seller side")
    row["Difference"] = _sum_col("Difference")
    return row


def _sort_matched(df: pd.DataFrame) -> pd.DataFrame:
    return df.sort_values(
        ["Item", "Invoice date — Buyer side"],
        ascending=[True, True],
        na_position="last",
    ).reset_index(drop=True)


def _sort_unmatched(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_sort_date"] = df.apply(
        lambda r: r["Invoice date — Seller side"]
        if pd.isna(r["Invoice date — Buyer side"])
        else r["Invoice date — Buyer side"],
        axis=1,
    )
    return (
        df.sort_values(["Item", "_sort_date"], ascending=[True, True], na_position="last")
        .drop(columns=["_sort_date"])
        .reset_index(drop=True)
    )


def build_result_table(match_result: MatchResult, buyer_short: str, seller_short: str) -> pd.DataFrame:
    """Build the combined result DataFrame with all three sections."""
    s1_rows = _matched_rows(match_result.exact_matches, buyer_short, seller_short)
    s1 = pd.DataFrame(s1_rows, columns=COLUMNS) if s1_rows else pd.DataFrame(columns=COLUMNS)
    if not s1.empty:
        s1 = _sort_matched(s1)

    s2_rows = _matched_rows(match_result.amount_mismatches, buyer_short, seller_short)
    s2 = pd.DataFrame(s2_rows, columns=COLUMNS) if s2_rows else pd.DataFrame(columns=COLUMNS)
    if not s2.empty:
        s2 = _sort_matched(s2)

    buy_rows = _unmatched_buy_rows(match_result.unmatched_buy, buyer_short, seller_short)
    sell_rows = _unmatched_sell_rows(match_result.unmatched_sell, buyer_short, seller_short)
    s3_rows = buy_rows + sell_rows
    s3 = pd.DataFrame(s3_rows, columns=COLUMNS) if s3_rows else pd.DataFrame(columns=COLUMNS)
    if not s3.empty:
        s3 = _sort_unmatched(s3)

    sep2 = pd.DataFrame([_blank_row(), _label_row(_LABEL_MISMATCH)], columns=COLUMNS)
    sep3 = pd.DataFrame([_blank_row(), _label_row(_LABEL_UNMATCHED)], columns=COLUMNS)

    def _with_subtotal(section: pd.DataFrame) -> list[pd.DataFrame]:
        if section.empty:
            return [section]
        return [section, pd.DataFrame([_subtotal_row(section)], columns=COLUMNS)]

    return pd.concat(
        [*_with_subtotal(s1), sep2, *_with_subtotal(s2), sep3, *_with_subtotal(s3)],
        ignore_index=True,
    )


def export_to_excel(result_df: pd.DataFrame, output_path: str) -> None:
    """Export the result DataFrame to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active

    translated_cols = [t(COL_KEY[col]) for col in COLUMNS]
    ws.append(translated_cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    col_indices = {col: idx + 1 for idx, col in enumerate(COLUMNS)}

    for _, row in result_df.iterrows():
        values = []
        for col in COLUMNS:
            val = None if pd.isna(row[col]) else row[col]
            if isinstance(val, str):
                if col == "Buyer Side" and val in SEP_KEY:
                    val = t(SEP_KEY[val])
                elif col == "Item" and val == _SUBTOTAL_ITEM_LABEL:
                    val = t("subtotal")
            values.append(val)
        ws.append(values)
        row_idx = ws.max_row

        buyer_val = row["Buyer Side"]
        item_val = row["Item"]
        is_separator = isinstance(buyer_val, str) and buyer_val in _SEPARATOR_LABELS
        is_subtotal = isinstance(item_val, str) and item_val == _SUBTOTAL_ITEM_LABEL

        if is_separator:
            for cell in ws[row_idx]:
                cell.fill = gray_fill
                cell.font = Font(bold=True, color="FF000000")
        elif is_subtotal:
            for cell in ws[row_idx]:
                cell.font = Font(bold=True, color="FF000000")

        for col in _DATE_COLS:
            cell = ws.cell(row=row_idx, column=col_indices[col])
            if cell.value is not None:
                cell.number_format = _DATE_FORMAT

        for col in _MONEY_COLS:
            cell = ws.cell(row=row_idx, column=col_indices[col])
            if cell.value is not None:
                cell.number_format = _MONEY_FORMAT

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(t(COL_KEY[col_name]))
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
